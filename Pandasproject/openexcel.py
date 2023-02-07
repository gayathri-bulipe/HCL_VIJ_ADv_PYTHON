import openpyxl
wb=openpyxl.load_workbook("..//data/Student.xlsx")
sheet=wb.active
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3).value
data1=sheet['A1:A10']
print(data)
print(data1)
print(data3)
#access all cells in row 1
print(sheet.max_row)
print(sheet.max_column)
#for row in sheet.rows:
#   print([data.value for data in row])
for i in range(2,12):
    print(sheet.cell(row=i,column=1).value)